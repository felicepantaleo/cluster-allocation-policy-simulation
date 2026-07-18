"""Map NGT working-package membership (budget xlsx) to cluster usernames.

    python tools/map_users_wp.py --xlsx <NGT_ALL.xlsx> --raw data/monit \
        --out data/derived/user_wp.json

Person rows are parsed from the WP1..WP4 and Mngmt sheets ("Surname
Firstname - Surname Firstname (DEPT)" in the details column; placeholder
rows are skipped). Each name is resolved to CERN account usernames via
anonymous LDAP on xldap.cern.ch (displayName token match, all tokens in
any order); the resolved usernames are intersected with the namespaces
observed on the cluster. A person in several WPs is assigned by priority
WP2 > WP3 > WP1 > WP4 > Management (PMC instruction: WP2 and WP3 win).

Output: {namespace: {"wp": ..., "name": ...}} plus an unresolved-names and
unmatched-namespaces report on stdout. The file contains usernames, so it
lives under data/ (gitignored).
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import unicodedata
from pathlib import Path

import openpyxl

WP_PRIORITY = {"WP2": 0, "WP3": 1, "WP1": 2, "WP4": 3, "Management": 4}
SHEETS = {"WP1": "WP1", "WP2": "WP2", "WP3": "WP3", "WP4": "WP4",
          "Mngmt": "Management"}
NAME_RE = re.compile(r"^([A-Za-zÀ-ž'. -]+?) - ")


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def parse_people(xlsx: str) -> dict[str, str]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    people: dict[str, str] = {}
    for sheet, wp in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            details = row[3] if sheet != "IT" else row[3]
            if not isinstance(details, str) or details.startswith(("[", "{")):
                continue
            m = NAME_RE.match(details.strip())
            if not m:
                continue
            name = " ".join(m.group(1).split())
            if name.lower() in ("overheads", "total"):
                continue
            cur = people.get(name)
            if cur is None or WP_PRIORITY[wp] < WP_PRIORITY[cur]:
                people[name] = wp
    return people


def ldap_usernames(name: str) -> list[str]:
    toks = [t for t in re.split(r"[ -]", strip_accents(name)) if len(t) > 1]
    filt = "(&(objectClass=user)" + "".join(
        f"(displayName=*{t}*)" for t in toks) + ")"
    out = subprocess.run(
        ["ldapsearch", "-x", "-H", "ldap://xldap.cern.ch", "-b",
         "OU=Users,OU=Organic Units,DC=cern,DC=ch", filt, "cn"],
        capture_output=True, text=True, timeout=30, check=False).stdout
    return [line.split(":", 1)[1].strip().lower()
            for line in out.splitlines() if line.startswith("cn:")]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--raw", default="data/monit")
    ap.add_argument("--out", default="data/derived/user_wp.json")
    args = ap.parse_args()

    people = parse_people(args.xlsx)
    print(f"{len(people)} people in WP sheets "
          f"({sum(1 for w in people.values() if w.startswith('WP'))} in WPs)")

    namespaces: set[str] = set()
    for f in sorted(Path(args.raw).glob("user_ns.*.json")):
        for s in json.loads(f.read_text()):
            namespaces.add(s["metric"]["namespace"])
    print(f"{len(namespaces)} user namespaces observed on the cluster")

    mapping: dict[str, dict] = {}
    unresolved: list[str] = []
    for name, wp in sorted(people.items()):
        try:
            cns = ldap_usernames(name)
        except Exception:
            cns = []
        hits = [c for c in cns if c in namespaces]
        if not cns:
            unresolved.append(name)
        for c in hits:
            prev = mapping.get(c)
            if prev is None or WP_PRIORITY[wp] < WP_PRIORITY[prev["wp"]]:
                mapping[c] = {"wp": wp, "name": name}

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(mapping, indent=2, sort_keys=True))

    by_wp: dict[str, int] = {}
    for v in mapping.values():
        by_wp[v["wp"]] = by_wp.get(v["wp"], 0) + 1
    print(f"mapped {len(mapping)} namespaces to WPs: {by_wp}")
    print(f"unresolved names ({len(unresolved)}): {unresolved}")
    unmatched = sorted(namespaces - set(mapping))
    print(f"namespaces without WP ({len(unmatched)}), first 30: {unmatched[:30]}")


if __name__ == "__main__":
    main()
